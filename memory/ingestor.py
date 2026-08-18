from pathlib import Path
from typing import Iterator
import PyPDF2
from docx import Document
from openpyxl import load_workbook


def extract_text_from_pdf(path: str) -> str:
    text = []
    with open(path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            text.append(page.extract_text() or "")
    return "\n".join(text)


def extract_text_from_docx(path: str) -> str:
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)


def extract_text_from_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def extract_text_from_xlsx(path: str) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            rows.append(" | ".join(str(c) if c is not None else "" for c in row))
    return "\n".join(rows)


def ingest_file(path: str) -> tuple[str, str]:
    path = Path(path)
    ext = path.suffix.lower()
    extractors = {
        ".pdf": extract_text_from_pdf,
        ".docx": extract_text_from_docx,
        ".txt": extract_text_from_txt,
        ".md": extract_text_from_txt,
        ".csv": extract_text_from_txt,
        ".xlsx": extract_text_from_xlsx,
    }
    extractor = extractors.get(ext)
    if not extractor:
        raise ValueError(f"Unsupported file type: {ext}")
    text = extractor(str(path))
    return text, ext


def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 200) -> list[str]:
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start += chunk_size - overlap
    return chunks


def ingest_and_chunk(path: str, chunk_size: int = 1000, overlap: int = 200) -> list[dict]:
    text, ext = ingest_file(path)
    chunks = chunk_text(text, chunk_size, overlap)
    return [
        {
            "text": chunk,
            "metadata": {"source": path, "type": ext, "chunk": i},
        }
        for i, chunk in enumerate(chunks)
    ]
